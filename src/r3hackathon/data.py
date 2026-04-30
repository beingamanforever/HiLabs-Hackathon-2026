from __future__ import annotations

from collections import Counter
import math
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .settings import BASE_DATA_FILE, CLAIMS_DATA_FILE, TELEHEALTH_POS_CODES


def _safe_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _pick_mode(counter: Counter) -> str | None:
    if not counter:
        return None
    return counter.most_common(1)[0][0]


def _counter_entropy(counter: Counter) -> float:
    total = sum(counter.values())
    if total <= 0:
        return 0.0
    entropy = 0.0
    for count in counter.values():
        probability = count / total
        entropy -= probability * math.log(probability, 2)
    return entropy


def load_base_data(base_path: Path | str = BASE_DATA_FILE) -> pd.DataFrame:
    """Load the base workbook.

    Row-count-agnostic: works on the local 2,493-row training set, the PS's
    1,500-record reference target, or any unseen holdout the judges hand us.
    If an `eval_set_flag` column exists (judge-injected for the holdout),
    we subset to those rows; otherwise we process the entire sheet.
    """
    # Try the named sheet first; fall back to the active sheet for holdouts
    # that may use a different sheet name.
    try:
        df = pd.read_excel(base_path, sheet_name="Base Data", header=1)
    except ValueError:
        df = pd.read_excel(base_path, header=1)

    if "eval_set_flag" in df.columns:
        before = len(df)
        df = df[df["eval_set_flag"].fillna(0).astype(int) == 1].copy()
        print(f"[load_base_data] eval_set_flag found: filtered {before} → {len(df)} rows")

    return df.reset_index(drop=True)


def aggregate_claims_for_base(
    base_df: pd.DataFrame,
    claims_path: Path | str = CLAIMS_DATA_FILE,
) -> pd.DataFrame:
    base_npis = {
        int(npi)
        for npi in pd.to_numeric(base_df["OrigNPI"], errors="coerce").dropna().tolist()
    }

    workbook = load_workbook(claims_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    header_index = {str(header): idx for idx, header in enumerate(headers) if header is not None}

    required_columns = {
        "CLAIMS_D_PRIMARY_HCP_NPI",
        "CLAIMS_PRIMARY_HCP_ZIP_5",
        "CLAIMS_PRIMARY_HCP_STATE",
        "CLAIMS_DATE_OF_SERVICE",
        "CLAIMS_D_PLACE_OF_SERVICE_CODE",
        "CLAIMS_SITE_OF_CARE_CATEGORY",
        "CLAIMS_AMOUNT_BILLED",
        "HCO_ATTR_HCO_NPI_ADDRESS_LINE_1",
        "HCO_ATTR_HCO_NPI_CITY",
        "HCO_ATTR_HCO_NPI_STATE",
        "HCO_ATTR_HCO_NPI_ZIP_5",
    }
    missing = sorted(required_columns - set(header_index))
    if missing:
        raise ValueError(f"Claims workbook is missing expected columns: {missing}")

    aggregates: dict[int, dict[str, Any]] = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        npi = _safe_int(row[header_index["CLAIMS_D_PRIMARY_HCP_NPI"]])
        if npi is None or npi not in base_npis:
            continue

        record = aggregates.setdefault(
            npi,
            {
                "claim_rows": 0,
                "amount_billed_sum": 0.0,
                "latest_dos": None,
                "claims_zip_counter": Counter(),
                "claims_state_counter": Counter(),
                "site_of_care_counter": Counter(),
                "claims_location_details": {},
                "telehealth_flag": 0,
                "hco_address_line_1": None,
                "hco_city": None,
                "hco_state": None,
                "hco_zip_5": None,
            },
        )

        record["claim_rows"] += 1

        billed = _safe_float(row[header_index["CLAIMS_AMOUNT_BILLED"]])
        if billed is not None:
            record["amount_billed_sum"] += billed

        dos = row[header_index["CLAIMS_DATE_OF_SERVICE"]]
        if dos is not None and (record["latest_dos"] is None or dos > record["latest_dos"]):
            record["latest_dos"] = dos

        claims_zip = _safe_int(row[header_index["CLAIMS_PRIMARY_HCP_ZIP_5"]])
        if claims_zip is not None:
            claims_zip_str = str(claims_zip)
            record["claims_zip_counter"][claims_zip_str] += 1
        else:
            claims_zip_str = None

        claims_state = row[header_index["CLAIMS_PRIMARY_HCP_STATE"]]
        if claims_state:
            claims_state_str = str(claims_state)
            record["claims_state_counter"][claims_state_str] += 1
        else:
            claims_state_str = None

        if claims_zip_str is not None:
            location_detail = record["claims_location_details"].setdefault(
                claims_zip_str,
                {
                    "count": 0,
                    "latest_dos": None,
                    "state_counter": Counter(),
                },
            )
            location_detail["count"] += 1
            if claims_state_str:
                location_detail["state_counter"][claims_state_str] += 1
            if dos is not None and (
                location_detail["latest_dos"] is None or dos > location_detail["latest_dos"]
            ):
                location_detail["latest_dos"] = dos

        site_of_care = row[header_index["CLAIMS_SITE_OF_CARE_CATEGORY"]]
        if site_of_care:
            record["site_of_care_counter"][str(site_of_care)] += 1

        pos_code = row[header_index["CLAIMS_D_PLACE_OF_SERVICE_CODE"]]
        if str(pos_code) in TELEHEALTH_POS_CODES:
            record["telehealth_flag"] = 1

        if record["hco_address_line_1"] is None:
            record["hco_address_line_1"] = row[header_index["HCO_ATTR_HCO_NPI_ADDRESS_LINE_1"]]
        if record["hco_city"] is None:
            record["hco_city"] = row[header_index["HCO_ATTR_HCO_NPI_CITY"]]
        if record["hco_state"] is None:
            record["hco_state"] = row[header_index["HCO_ATTR_HCO_NPI_STATE"]]
        if record["hco_zip_5"] is None:
            hco_zip = _safe_int(row[header_index["HCO_ATTR_HCO_NPI_ZIP_5"]])
            record["hco_zip_5"] = str(hco_zip) if hco_zip is not None else None

    rows = []
    for npi, record in aggregates.items():
        claim_rows = record["claim_rows"]
        sorted_locations = sorted(
            record["claims_location_details"].items(),
            key=lambda item: (item[1]["count"], item[1]["latest_dos"] or pd.Timestamp.min),
            reverse=True,
        )
        dominant_zip = sorted_locations[0][0] if sorted_locations else None
        dominant_detail = sorted_locations[0][1] if sorted_locations else None
        secondary_zip = sorted_locations[1][0] if len(sorted_locations) > 1 else None
        secondary_detail = sorted_locations[1][1] if len(sorted_locations) > 1 else None

        rows.append(
            {
                "OrigNPI": npi,
                "claim_rows": claim_rows,
                "amount_billed_sum": record["amount_billed_sum"],
                "amount_billed_mean": record["amount_billed_sum"] / claim_rows if claim_rows else 0.0,
                "latest_dos": record["latest_dos"],
                "claims_zip": _pick_mode(record["claims_zip_counter"]),
                "claims_state": _pick_mode(record["claims_state_counter"]),
                "site_of_care_mode": _pick_mode(record["site_of_care_counter"]),
                "n_claim_zip_locations": len(record["claims_location_details"]),
                "claims_location_entropy": _counter_entropy(record["claims_zip_counter"]),
                "dominant_claims_zip": dominant_zip,
                "dominant_claims_state": _pick_mode(dominant_detail["state_counter"]) if dominant_detail else None,
                "dominant_claim_share": (dominant_detail["count"] / claim_rows) if dominant_detail and claim_rows else 0.0,
                "dominant_claim_latest_dos": dominant_detail["latest_dos"] if dominant_detail else None,
                "secondary_claims_zip": secondary_zip,
                "secondary_claims_state": _pick_mode(secondary_detail["state_counter"]) if secondary_detail else None,
                "secondary_claim_share": (secondary_detail["count"] / claim_rows) if secondary_detail and claim_rows else 0.0,
                "secondary_claim_latest_dos": secondary_detail["latest_dos"] if secondary_detail else None,
                "telehealth_flag": int(record["telehealth_flag"]),
                "hco_address_line_1": record["hco_address_line_1"],
                "hco_city": record["hco_city"],
                "hco_state": record["hco_state"],
                "hco_zip_5": record["hco_zip_5"],
            }
        )

    return pd.DataFrame(rows)
